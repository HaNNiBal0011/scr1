"""
Экспорт результатов в различные форматы
"""

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Any
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from scraper.base_scraper import ScrapingResult, ScrapingStatus

class ExportManager:
    """Менеджер экспорта результатов"""
    
    def __init__(self, output_dir: str = "exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def export_to_csv(self, results: List[ScrapingResult], file_path: str, 
                     options: Optional[Dict] = None) -> bool:
        """Экспорт в CSV формат"""
        try:
            if options is None:
                options = {
                    'fields': ['id', 'site', 'name', 'price', 'availability', 'status'],
                    'include_headers': True,
                    'success_only': False
                }
            
            # Фильтрация результатов
            filtered_results = self._filter_results(results, options.get('success_only', False))
            
            if not filtered_results:
                self.logger.warning("Нет данных для экспорта")
                return False
            
            # Подготовка данных
            data_rows = []
            
            # Заголовки
            if options.get('include_headers', True):
                headers = self._get_field_headers(options.get('fields', []))
                data_rows.append(headers)
            
            # Данные
            for result in filtered_results:
                row = self._extract_row_data(result, options.get('fields', []))
                data_rows.append(row)
            
            # Запись в файл
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                writer.writerows(data_rows)
            
            self.logger.info(f"CSV экспорт завершен: {file_path} ({len(filtered_results)} записей)")
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка экспорта в CSV: {e}")
            return False
    
    def export_to_json(self, results: List[ScrapingResult], file_path: str,
                      options: Optional[Dict] = None) -> bool:
        """Экспорт в JSON формат"""
        try:
            if options is None:
                options = {
                    'fields': ['id', 'site', 'name', 'price', 'availability', 'status'],
                    'success_only': False,
                    'pretty': True
                }
            
            # Фильтрация результатов
            filtered_results = self._filter_results(results, options.get('success_only', False))
            
            if not filtered_results:
                self.logger.warning("Нет данных для экспорта")
                return False
            
            # Подготовка данных
            export_data = {
                'metadata': {
                    'export_date': datetime.now().isoformat(),
                    'total_records': len(filtered_results),
                    'success_count': sum(1 for r in filtered_results if r.status == ScrapingStatus.SUCCESS),
                    'error_count': sum(1 for r in filtered_results if r.status != ScrapingStatus.SUCCESS),
                    'fields_exported': options.get('fields', [])
                },
                'results': []
            }
            
            # Данные результатов
            for result in filtered_results:
                result_data = self._extract_result_data(result, options.get('fields', []))
                export_data['results'].append(result_data)
            
            # Запись в файл
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                if options.get('pretty', True):
                    json.dump(export_data, jsonfile, indent=2, ensure_ascii=False)
                else:
                    json.dump(export_data, jsonfile, ensure_ascii=False)
            
            self.logger.info(f"JSON экспорт завершен: {file_path} ({len(filtered_results)} записей)")
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка экспорта в JSON: {e}")
            return False
    
    def export_to_excel(self, results: List[ScrapingResult], file_path: str,
                       options: Optional[Dict] = None) -> bool:
        """Экспорт в Excel формат"""
        if not OPENPYXL_AVAILABLE:
            self.logger.error("openpyxl не установлен. Экспорт в Excel недоступен.")
            return False
        
        try:
            if options is None:
                options = {
                    'fields': ['id', 'site', 'name', 'price', 'availability', 'status'],
                    'include_headers': True,
                    'success_only': False,
                    'styling': True
                }
            
            # Фильтрация результатов
            filtered_results = self._filter_results(results, options.get('success_only', False))
            
            if not filtered_results:
                self.logger.warning("Нет данных для экспорта")
                return False
            
            # Создание рабочей книги
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Результаты скрейпинга"
            
            # Заголовки
            if options.get('include_headers', True):
                headers = self._get_field_headers(options.get('fields', []))
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    
                    if options.get('styling', True):
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
            
            # Данные
            for row_idx, result in enumerate(filtered_results, 2):
                row_data = self._extract_row_data(result, options.get('fields', []))
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    if options.get('styling', True):
                        # Стилизация в зависимости от статуса
                        if 'status' in options.get('fields', []):
                            status_col = options.get('fields', []).index('status') + 1
                            if col_idx == status_col:
                                if result.status == ScrapingStatus.SUCCESS:
                                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                                    cell.font = Font(color="155724")
                                else:
                                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                                    cell.font = Font(color="721C24")
            
            # Автоширина колонок
            if options.get('styling', True):
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохранение
            wb.save(file_path)
            
            self.logger.info(f"Excel экспорт завершен: {file_path} ({len(filtered_results)} записей)")
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка экспорта в Excel: {e}")
            return False
    
    def export_to_html(self, results: List[ScrapingResult], file_path: str,
                      options: Optional[Dict] = None) -> bool:
        """Экспорт в HTML формат"""
        try:
            if options is None:
                options = {
                    'fields': ['id', 'site', 'name', 'price', 'availability', 'status'],
                    'include_headers': True,
                    'success_only': False,
                    'styling': True
                }
            
            # Фильтрация результатов
            filtered_results = self._filter_results(results, options.get('success_only', False))
            
            if not filtered_results:
                self.logger.warning("Нет данных для экспорта")
                return False
            
            # Генерация HTML
            html_content = self._generate_html_report(filtered_results, options)
            
            # Запись в файл
            with open(file_path, 'w', encoding='utf-8') as htmlfile:
                htmlfile.write(html_content)
            
            self.logger.info(f"HTML экспорт завершен: {file_path} ({len(filtered_results)} записей)")
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка экспорта в HTML: {e}")
            return False
    
    def _filter_results(self, results: List[ScrapingResult], success_only: bool) -> List[ScrapingResult]:
        """Фильтрация результатов"""
        if success_only:
            return [r for r in results if r.status == ScrapingStatus.SUCCESS]
        return results
    
    def _get_field_headers(self, fields: List[str]) -> List[str]:
        """Получение заголовков полей"""
        field_mapping = {
            'id': 'ID товара',
            'site': 'Сайт',
            'name': 'Название',
            'price': 'Цена',
            'old_price': 'Старая цена',
            'availability': 'Наличие',
            'url': 'URL',
            'image_url': 'URL изображения',
            'description': 'Описание',
            'method_used': 'Метод',
            'response_time': 'Время ответа (сек)',
            'status': 'Статус',
            'error_message': 'Ошибка',
            'attempts': 'Попытки'
        }
        
        return [field_mapping.get(field, field) for field in fields]
    
    def _extract_row_data(self, result: ScrapingResult, fields: List[str]) -> List[Any]:
        """Извлечение данных строки"""
        row = []
        
        for field in fields:
            if field == 'id':
                row.append(result.product.id if result.product else '')
            elif field == 'site':
                row.append(result.product.site if result.product else '')
            elif field == 'name':
                row.append(result.product.name if result.product else '')
            elif field == 'price':
                if result.product and result.product.price is not None:
                    row.append(f"{result.product.price:.2f}")
                else:
                    row.append('')
            elif field == 'old_price':
                if result.product and result.product.old_price is not None:
                    row.append(f"{result.product.old_price:.2f}")
                else:
                    row.append('')
            elif field == 'availability':
                row.append(result.product.availability if result.product else '')
            elif field == 'url':
                row.append(result.product.url if result.product else '')
            elif field == 'image_url':
                row.append(result.product.image_url if result.product else '')
            elif field == 'description':
                row.append(result.product.description if result.product else '')
            elif field == 'method_used':
                row.append(result.method_used.value if result.method_used else '')
            elif field == 'response_time':
                row.append(f"{result.response_time:.2f}")
            elif field == 'status':
                row.append('Успешно' if result.status == ScrapingStatus.SUCCESS else 'Ошибка')
            elif field == 'error_message':
                row.append(result.error_message)
            elif field == 'attempts':
                row.append(result.attempts)
            else:
                row.append('')
        
        return row
    
    def _extract_result_data(self, result: ScrapingResult, fields: List[str]) -> Dict[str, Any]:
        """Извлечение данных результата для JSON"""
        data = {}
        
        for field in fields:
            if field == 'id':
                data[field] = result.product.id if result.product else None
            elif field == 'site':
                data[field] = result.product.site if result.product else None
            elif field == 'name':
                data[field] = result.product.name if result.product else None
            elif field == 'price':
                data[field] = result.product.price if result.product else None
            elif field == 'old_price':
                data[field] = result.product.old_price if result.product else None
            elif field == 'availability':
                data[field] = result.product.availability if result.product else None
            elif field == 'url':
                data[field] = result.product.url if result.product else None
            elif field == 'image_url':
                data[field] = result.product.image_url if result.product else None
            elif field == 'description':
                data[field] = result.product.description if result.product else None
            elif field == 'method_used':
                data[field] = result.method_used.value if result.method_used else None
            elif field == 'response_time':
                data[field] = result.response_time
            elif field == 'status':
                data[field] = result.status.value
            elif field == 'error_message':
                data[field] = result.error_message
            elif field == 'attempts':
                data[field] = result.attempts
            else:
                data[field] = None
        
        return data
    
    def _generate_html_report(self, results: List[ScrapingResult], options: Dict) -> str:
        """Генерация HTML отчета"""
        fields = options.get('fields', [])
        headers = self._get_field_headers(fields)
        styling = options.get('styling', True)
        
        # CSS стили
        css_styles = """
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            h1 { color: #333; border-bottom: 2px solid #366092; padding-bottom: 10px; }
            .summary { background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin-bottom: 20px; }
            .summary div { display: inline-block; margin-right: 30px; }
            table { border-collapse: collapse; width: 100%; margin-top: 20px; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            th { background-color: #366092; color: white; font-weight: bold; }
            tr:nth-child(even) { background-color: #f2f2f2; }
            .success { background-color: #d4edda; color: #155724; }
            .error { background-color: #f8d7da; color: #721c24; }
            .timestamp { color: #666; font-size: 0.9em; }
        </style>
        """ if styling else ""
        
        # Статистика
        total_count = len(results)
        success_count = sum(1 for r in results if r.status == ScrapingStatus.SUCCESS)
        error_count = total_count - success_count
        
        # Формирование HTML
        html = f"""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Отчет о скрейпинге товаров</title>
            {css_styles}
        </head>
        <body>
            <h1>Отчет о скрейпинге товаров</h1>
            
            <div class="summary">
                <div><strong>Дата создания:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
                <div><strong>Всего товаров:</strong> {total_count}</div>
                <div><strong>Успешно:</strong> {success_count}</div>
                <div><strong>Ошибок:</strong> {error_count}</div>
                <div><strong>Успешность:</strong> {(success_count/total_count*100):.1f}%</div>
            </div>
            
            <table>
                <thead>
                    <tr>
        """
        
        # Заголовки таблицы
        for header in headers:
            html += f"<th>{header}</th>"
        
        html += """
                    </tr>
                </thead>
                <tbody>
        """
        
        # Данные таблицы
        for result in results:
            row_data = self._extract_row_data(result, fields)
            status_class = "success" if result.status == ScrapingStatus.SUCCESS else "error"
            
            html += f'<tr class="{status_class}">'
            for value in row_data:
                html += f"<td>{value}</td>"
            html += "</tr>"
        
        html += """
                </tbody>
            </table>
            
            <div class="timestamp">
                <p>Отчет создан: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def export_statistics(self, results: List[ScrapingResult], file_path: str) -> bool:
        """Экспорт статистики"""
        try:
            total_count = len(results)
            success_count = sum(1 for r in results if r.status == ScrapingStatus.SUCCESS)
            error_count = total_count - success_count
            
            # Статистика по сайтам
            site_stats = {}
            method_stats = {}
            
            for result in results:
                if result.product:
                    site = result.product.site
                    site_stats[site] = site_stats.get(site, {'success': 0, 'error': 0})
                    
                    if result.status == ScrapingStatus.SUCCESS:
                        site_stats[site]['success'] += 1
                    else:
                        site_stats[site]['error'] += 1
                
                if result.method_used:
                    method = result.method_used.value
                    method_stats[method] = method_stats.get(method, {'success': 0, 'error': 0})
                    
                    if result.status == ScrapingStatus.SUCCESS:
                        method_stats[method]['success'] += 1
                    else:
                        method_stats[method]['error'] += 1
            
            # Среднее время ответа
            response_times = [r.response_time for r in results if r.response_time > 0]
            avg_response_time = sum(response_times) / len(response_times) if response_times else 0
            
            # Формирование статистики
            statistics = {
                'summary': {
                    'total_products': total_count,
                    'successful': success_count,
                    'errors': error_count,
                    'success_rate': (success_count / total_count * 100) if total_count > 0 else 0,
                    'average_response_time': avg_response_time,
                    'export_date': datetime.now().isoformat()
                },
                'site_statistics': site_stats,
                'method_statistics': method_stats,
                'error_analysis': self._analyze_errors(results)
            }
            
            # Сохранение в JSON
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(statistics, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"Статистика экспортирована: {file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка экспорта статистики: {e}")
            return False
    
    def _analyze_errors(self, results: List[ScrapingResult]) -> Dict[str, int]:
        """Анализ ошибок"""
        error_types = {}
        
        for result in results:
            if result.status != ScrapingStatus.SUCCESS and result.error_message:
                error_msg = result.error_message.lower()
                
                # Классификация ошибок
                if 'timeout' in error_msg:
                    error_type = 'Timeout'
                elif 'connection' in error_msg:
                    error_type = 'Connection Error'
                elif 'blocked' in error_msg or 'captcha' in error_msg:
                    error_type = 'Site Blocking'
                elif 'selenium' in error_msg or 'webdriver' in error_msg:
                    error_type = 'WebDriver Error'
                elif 'parse' in error_msg or 'parsing' in error_msg:
                    error_type = 'Parsing Error'
                else:
                    error_type = 'Other Error'
                
                error_types[error_type] = error_types.get(error_type, 0) + 1
        
        return error_types
    
    def create_export_template(self, format_type: str, file_path: str) -> bool:
        """Создание шаблона для экспорта"""
        try:
            if format_type.lower() == 'csv':
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['ID товара', 'Сайт', 'Название', 'Цена', 'Наличие', 'Статус'])
                    writer.writerow(['123456', 'rozetka', 'Пример товара', '1000.00', 'В наличии', 'Успешно'])
            
            elif format_type.lower() == 'json':
                template = {
                    'metadata': {
                        'export_date': 'YYYY-MM-DDTHH:MM:SS',
                        'total_records': 1,
                        'success_count': 1,
                        'error_count': 0
                    },
                    'results': [
                        {
                            'id': '123456',
                            'site': 'rozetka',
                            'name': 'Пример товара',
                            'price': 1000.00,
                            'availability': 'В наличии',
                            'status': 'success'
                        }
                    ]
                }
                
                with open(file_path, 'w', encoding='utf-8') as jsonfile:
                    json.dump(template, jsonfile, indent=2, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка создания шаблона: {e}")
            return False
